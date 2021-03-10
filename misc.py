#! python
# -*- coding: utf-8 -*-
"""Editor's collection of wxpj miscellaneous functions

Author: Kazuya O'moto <komoto@jeol.co.jp>
"""
from __future__ import (division, print_function,
                        absolute_import, unicode_literals)
from datetime import timedelta
import openpyxl as pxl
import configparser
import time
import sys
import os
import wx
import numpy as np

LITERAL_TYPE = str if sys.version_info >= (3,0) else (str,unicode)


def xlread(worksheet, r=1, c=1):
    """Read data from worksheet starting from cell(r,c)"""
    ## values = list(worksheet.values)
    ## return np.array([[eval(x) for x in row[c-1:]] for row in values[r-1:]])
    try:
        data = []
        for row in worksheet.iter_rows(min_row=r, min_col=c):
            data.append([cell.value for cell in row])
    finally:
        return np.array(data)


def xlwrite(worksheet, data, r=1, c=1):
    """Write data to worksheet starting from cell(r,c)"""
    for j,ln in enumerate(data):
        for k,x in enumerate(ln):
            cell = worksheet.cell(j+r, k+c)
            cell.value = x


def _eval_array(x):
    try:
        return eval(x)
    except Exception:
        return x

def _format_array(arr):
    return (repr(arr)
            .replace('array', 'np.array')
            .replace('       [', '[') # erase (7) phantom of "array(["
            .replace('([[', '([\n[')
            .replace(']])', '],\n])')
            )

class ConfigData(object):
    """Config paraser wraps for sharing data among plugins
    """
    def __init__(self, path, section):
        self.parser = configparser.ConfigParser()
        ## self.parser.optionxform = str # 大文字／小文字の区別
        self.data = {}
        self.path = path
        self.section = section
        self.load(verbose=0)
    
    def __getitem__(self, k):
        return self.data[k]
    
    def __setitem__(self, k, v):
        self.data[k] = v
    
    def load(self, keys=None, verbose=True):
        """Load configurations of the given `section
        if `keys None, load ALL keys from the section
        """
        if verbose:
            if wx.MessageBox("Do you want to reload configration?\n"
                "\n The current data is to be overwritten with the original data:"
                "\n {!r} [{}], keys={!r}".format(self.path, self.section, keys), "Load",
                style=wx.YES_NO|wx.ICON_INFORMATION) != wx.YES:
                    return
        
        if isinstance(keys, LITERAL_TYPE):
            keys = [keys]
        
        self.parser.read(self.path)
        entry = self.parser[self.section]
        for t in keys or entry:
            self.data[t] = _eval_array(entry[t])
    ## load_config = load
    
    def save(self, keys=None, verbose=True):
        """Save configurations of the given `section
        if `keys None, save ALL keys into the section
        """
        if verbose:
            if wx.MessageBox("Do you want to save configration?\n"
                "\n The current data is to be saved and overwrites the original data:"
                "\n {!r} [{}], keys={!r}".format(self.path, self.section, keys), "Save",
                style=wx.YES_NO|wx.ICON_INFORMATION) != wx.YES:
                    return
        
        if isinstance(keys, LITERAL_TYPE):
            keys = [keys]
        try:
            opt = np.get_printoptions()
            np.set_printoptions(linewidth=256, threshold=np.inf) # printing all(inf) elements
            ## np.set_printoptions(formatter={'float':'{:-12.8g}'.format})
            
            entry = self.parser[self.section]
            for t in keys or entry:
                if t in self.parser.defaults():
                    continue
                entry[t] = _format_array(self.data[t])
            self.parser.write(open(self.path, 'w'))
        finally:
            np.set_printoptions(**opt)
    ## save_config = save
    
    def export(self, keys, r=3, c=3, verbose=True):
        """Export configurations of the given `section
        """
        ## xlpath = "config-report.xlsx"
        name, _ext = os.path.splitext(os.path.basename(self.path))
        xlpath = "config-report-{}.xlsx".format(name)
        if verbose:
            if wx.MessageBox("Exporting configration to excel file {!r}".format(xlpath),
                style=wx.YES_NO|wx.ICON_INFORMATION) != wx.YES:
                    return
        try:
            wbook = pxl.load_workbook(xlpath)
            for key in keys:
                xlwrite(wbook[key], self.data[key], r, c)
            wbook.save(xlpath)
            return True
        
        except PermissionError as e:
            if wx.MessageBox("Please close {!r}. Press [OK] to continue.".format(xlpath),
                caption=str(e), style=wx.OK|wx.CANCEL|wx.ICON_WARNING) == wx.OK:
                    return self.export(keys, r, c, verbose=0)


def count_down(seconds):
    for d in range(int(seconds),0,-1):
        s = "wait.. {}".format(timedelta(seconds=d))
        print("\b"*80, s[:80], sep='', end='')
        sys.stdout.flush()
        time.sleep(1)
    print("")


class progress_counter(object):
    def __init__(self, max_count=None):
        self.max_count = max_count # None means endless count
        self.duration = 0
        self.count = 0
        self.tm_start = time.time()
    
    def __call__(self, message=''):
        self.count += 1
        self.duration = time.time() - self.tm_start
        dt = timedelta(seconds=int(self.duration))
        if self.max_count:
            p = self.count / self.max_count
            s = " {:2.0%} elapses {}".format(p, dt)
        else:
            p = 0
            s = " elapses {}".format(dt)
        s += ' ' + message
        print("\b" * 80, s[:80], sep='', end='')
        sys.stdout.flush()
        return p < 1


if __name__ == "__main__":
    import mwx
    
    config = ConfigData("pylots/config.ini", section='TEM')
    ## config.export('beamshift')
    mwx.deb()
    
    print(time.strftime("%Y/%m/%d %H:%M:%S %z"))
    count_down(3)
    
    count = progress_counter(10)
    print("progressing...", end='')
    while count():
        time.sleep(1)
    print()
